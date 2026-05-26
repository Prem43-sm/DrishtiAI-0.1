import calendar
import json
import os
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from PySide6.QtCore import QDate
from PySide6.QtWidgets import (
    QComboBox,
    QDateEdit,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from core.project_paths import ATTENDANCE_DIR, KNOWN_FACES_DIR, TIMETABLE_DIR, ensure_runtime_layout
from features.student_records_db import list_student_records


class AttendancePage(QWidget):
    def __init__(self):
        super().__init__()

        ensure_runtime_layout()
        self.base_dir = str(ATTENDANCE_DIR)
        self.face_dir = str(KNOWN_FACES_DIR)

        os.makedirs(self.base_dir, exist_ok=True)
        os.makedirs(self.face_dir, exist_ok=True)

        self.current_view_df = pd.DataFrame()

        main_layout = QVBoxLayout()

        title = QLabel("Attendance")
        title.setStyleSheet("font-size:18px; font-weight:bold;")
        main_layout.addWidget(title)

        top_bar = QHBoxLayout()
        self.class_selector = QComboBox()
        self.load_classes()

        self.date_picker = QDateEdit()
        self.date_picker.setCalendarPopup(True)
        self.date_picker.setDisplayFormat("dd-MM-yyyy")
        self.date_picker.setDate(QDate.currentDate())

        load_btn = QPushButton("Load Date")
        load_btn.clicked.connect(self.load_attendance)

        export_btn = QPushButton("Export Excel")
        export_btn.clicked.connect(self.export_excel)

        class_month_export_btn = QPushButton("Export Class Month")
        class_month_export_btn.clicked.connect(self.export_class_month_report)

        top_bar.addWidget(QLabel("Class:"))
        top_bar.addWidget(self.class_selector)
        top_bar.addWidget(QLabel("Date:"))
        top_bar.addWidget(self.date_picker)
        top_bar.addWidget(load_btn)
        top_bar.addStretch()
        top_bar.addWidget(export_btn)
        top_bar.addWidget(class_month_export_btn)
        main_layout.addLayout(top_bar)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by name...")
        self.search_input.textChanged.connect(self.filter_table)
        main_layout.addWidget(self.search_input)

        self.table = QTableWidget()
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        main_layout.addWidget(self.table)

        self.setLayout(main_layout)
        self.active_class = ""
        self.active_period = None

    # ---------------------------------------------------------
    # CLASS LIST
    # ---------------------------------------------------------
    def load_classes(self):
        current = self.class_selector.currentText()
        self.class_selector.clear()

        classes = []
        if os.path.exists(self.face_dir):
            for cls in sorted(os.listdir(self.face_dir)):
                full = os.path.join(self.face_dir, cls)
                if os.path.isdir(full):
                    classes.append(cls)
                    self.class_selector.addItem(cls)

        if current and current in classes:
            self.class_selector.setCurrentText(current)

    # ---------------------------------------------------------
    # SINGLE DATE VIEW (FULL CLASS)
    # ---------------------------------------------------------
    def load_attendance(self):
        cls = self.class_selector.currentText().strip()
        if not cls:
            QMessageBox.warning(self, "Error", "Select a class first")
            return

        date_str = self.date_picker.date().toString("yyyy-MM-dd")
        view_df = self._build_single_date_view(cls, date_str)
        self.current_view_df = view_df
        self._render_table(view_df)

    def _get_students_for_class(self, cls):
        students = set()
        for row in self._student_details_for_class(cls).values():
            if str(row.get("status", "Active")).lower() != "left":
                name = str(row.get("student_name", "")).strip()
                if name:
                    students.add(name)
        class_dir = os.path.join(self.face_dir, cls)

        if os.path.exists(class_dir):
            for file_name in os.listdir(class_dir):
                if file_name.endswith(".npy"):
                    students.add(os.path.splitext(file_name)[0])

        # fallback from attendance files if needed
        class_att_dir = os.path.join(self.base_dir, cls)
        if os.path.exists(class_att_dir):
            for file_name in os.listdir(class_att_dir):
                if not file_name.endswith(".csv"):
                    continue
                try:
                    tmp = pd.read_csv(os.path.join(class_att_dir, file_name))
                    if "Name" in tmp.columns:
                        for n in tmp["Name"].dropna().astype(str):
                            if n.strip():
                                students.add(n.strip())
                except Exception:
                    continue

        return sorted(students)

    def _get_periods_for_date(self, cls, date_str):
        periods = set()

        # 1) Prefer timetable day periods
        timetable_file = os.path.join(str(TIMETABLE_DIR), f"{cls}.json")
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d")
            day_name = d.strftime("%A")
        except Exception:
            day_name = None

        if day_name and os.path.exists(timetable_file):
            try:
                with open(timetable_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                for p in data.get("days", {}).get(day_name, []):
                    period = p.get("period")
                    if period is not None:
                        periods.add(str(period))
            except Exception:
                pass

        # 2) Merge file-based periods for that exact date
        class_path = os.path.join(self.base_dir, cls)
        if os.path.exists(class_path):
            for file_name in os.listdir(class_path):
                if not file_name.startswith(f"{date_str}_P") or not file_name.endswith(".csv"):
                    continue
                period = file_name.split("_P")[-1].replace(".csv", "").strip()
                if period:
                    periods.add(period)

        if not periods:
            periods = {"1"}

        return sorted(periods, key=lambda x: int(x) if str(x).isdigit() else 9999)

    def _build_single_date_view(self, cls, date_str):
        students = self._get_students_for_class(cls)
        periods = self._get_periods_for_date(cls, date_str)

        rows = []
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            selected_date = datetime.now().date()

        today = datetime.now().date()
        is_sunday = selected_date.weekday() == 6
        is_future = selected_date > today

        for name in students:
            row = {"Name": name, "Class": cls}
            for p in periods:
                col = f"P{p}"
                if is_future:
                    row[col] = ""
                elif is_sunday:
                    row[col] = "H"
                else:
                    row[col] = "A"
            rows.append(row)

        df = pd.DataFrame(rows, columns=["Name", "Class"] + [f"P{p}" for p in periods])

        # fill present values
        class_path = os.path.join(self.base_dir, cls)
        for p in periods:
            file_path = os.path.join(class_path, f"{date_str}_P{p}.csv")
            if not os.path.exists(file_path):
                continue
            try:
                day_df = pd.read_csv(file_path)
            except Exception:
                continue

            if "Name" not in day_df.columns:
                continue

            for _, r in day_df.iterrows():
                n = str(r.get("Name", "")).strip()
                if not n:
                    continue
                if n not in set(df["Name"].astype(str)):
                    # include unknown/new name too
                    new_row = {"Name": n, "Class": cls}
                    for pp in periods:
                        col = f"P{pp}"
                        if is_future:
                            new_row[col] = ""
                        elif is_sunday:
                            new_row[col] = "H"
                        else:
                            new_row[col] = "A"
                    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

                t = str(r.get("Time", "")).strip()
                mark = f"P ({t})" if t and t.lower() != "nan" else "P"
                df.loc[df["Name"].astype(str) == n, f"P{p}"] = mark

        if not df.empty:
            df = df.sort_values("Name").reset_index(drop=True)
        return df

    def _render_table(self, df):
        if df.empty:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        self.table.setRowCount(len(df))
        self.table.setColumnCount(len(df.columns))
        self.table.setHorizontalHeaderLabels([str(c) for c in df.columns])

        for r in range(len(df)):
            for c in range(len(df.columns)):
                self.table.setItem(r, c, QTableWidgetItem(str(df.iat[r, c])))

    def filter_table(self):
        text = self.search_input.text().lower().strip()
        if self.current_view_df.empty:
            return
        if not text:
            self._render_table(self.current_view_df)
            return

        filtered = self.current_view_df[
            self.current_view_df["Name"].astype(str).str.lower().str.contains(text)
        ]
        self._render_table(filtered)

    def set_active_class(self, class_name, period=None):
        self.active_class = str(class_name or "").strip()
        self.active_period = period
        if not self.active_class:
            return

        if self.class_selector.findText(self.active_class) == -1:
            self.class_selector.addItem(self.active_class)
        self.class_selector.setCurrentText(self.active_class)

    def stop_auto_attendance(self):
        self.active_period = None

    # ---------------------------------------------------------
    # MONTHLY EXPORT (kept)
    # ---------------------------------------------------------
    def export_class_month_report(self):
        cls = self.class_selector.currentText().strip()
        if not cls:
            QMessageBox.warning(self, "Error", "Select a class first")
            return

        selected_date = self.date_picker.date()
        year = int(selected_date.year())
        month = int(selected_date.month())
        class_name, semester = self._split_class_and_semester(cls)
        month_name = datetime(year, month, 1).strftime("%B")
        output_dir = Path(self.base_dir) / f"{year}_attendance"
        output_dir.mkdir(parents=True, exist_ok=True)
        file_name = f"{self._safe_filename(class_name)}_{self._safe_filename(semester)}_{month_name}.xlsx"
        output_path = output_dir / file_name

        report_rows, date_periods = self._build_class_month_report_rows(cls, class_name, semester, year, month)
        if not report_rows:
            QMessageBox.warning(self, "Error", "No students found for selected class")
            return

        self._write_class_month_workbook(output_path, report_rows, date_periods)
        QMessageBox.information(self, "Done", f"Class monthly attendance saved:\n{output_path}")

    def export_excel(self):
        cls = self.class_selector.currentText().strip()
        if not cls:
            QMessageBox.warning(self, "Error", "Select a class first")
            return

        class_path = os.path.join(self.base_dir, cls)
        if not os.path.exists(class_path):
            QMessageBox.warning(self, "Error", "No attendance folder for selected class")
            return

        monthly_groups = self._load_monthly_data(class_path, cls)
        if not monthly_groups:
            QMessageBox.warning(self, "Error", "No attendance data to export")
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Monthly Report",
            f"{cls}_monthly_report.xlsx",
            "Excel (*.xlsx)",
        )
        if not path:
            return

        timetable_periods = self._get_timetable_periods(cls)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for month_key, month_df in sorted(monthly_groups.items()):
                students_df = self._get_all_students_for_class(cls, month_df)
                report_df = self._build_month_matrix(month_df, month_key, timetable_periods)
                report_df = self._ensure_all_students(report_df, students_df)
                if report_df.empty:
                    continue
                sheet_name = month_key.replace("-", "_")
                report_df.to_excel(writer, sheet_name=sheet_name, index=True)

        QMessageBox.information(self, "Done", "Monthly report exported successfully")

    def _build_class_month_report_rows(self, cls, class_name, semester, year, month):
        days_in_month = calendar.monthrange(year, month)[1]
        today = datetime.now().date()
        students = self._get_students_for_class(cls)
        student_details = self._student_details_for_class(cls)
        timetable = self._load_timetable(cls)
        date_periods = []
        rows = []

        for day in range(1, days_in_month + 1):
            date_obj = datetime(year, month, day).date()
            date_str = date_obj.strftime("%Y-%m-%d")
            day_periods = self._get_period_details_for_date(cls, date_str, timetable)
            for period_info in day_periods:
                date_periods.append(
                    {
                        "date": date_str,
                        "display_date": date_obj.strftime("%d-%m-%Y"),
                        "period": str(period_info.get("period", "")),
                        "subject": str(period_info.get("subject", "")).strip(),
                    }
                )

        marks = self._load_month_attendance_marks(cls, year, month)
        for index, name in enumerate(students, start=1):
            student_marks = []
            for item in date_periods:
                date_obj = datetime.strptime(item["date"], "%Y-%m-%d").date()
                key = (name, item["date"], item["period"])
                if key in marks:
                    student_marks.append(marks[key])
                elif date_obj > today:
                    student_marks.append("")
                elif date_obj.weekday() == 6:
                    student_marks.append("H")
                else:
                    student_marks.append("A")

            rows.append(
                {
                    "Batch ( year )": year,
                    "Class name": class_name,
                    "Semester": semester,
                    "Roll no.": student_details.get(name, {}).get("roll_number", ""),
                    "Name": name,
                    "contact no.": student_details.get(name, {}).get("contact_number", ""),
                    "marks": student_marks,
                }
            )

        return rows, date_periods

    def _student_details_for_class(self, cls):
        details = {}
        try:
            records = list_student_records(cls)
        except Exception:
            records = []
        for row in records:
            name = str(row.get("student_name", "")).strip()
            if name:
                details[name] = row
        return details

    def _load_month_attendance_marks(self, cls, year, month):
        marks = {}
        class_path = Path(self.base_dir) / cls
        if not class_path.exists():
            return marks

        prefix = f"{year:04d}-{month:02d}-"
        for csv_file in sorted(class_path.glob(f"{prefix}*_P*.csv")):
            date_part = csv_file.name.split("_P", 1)[0]
            period = csv_file.stem.split("_P")[-1].strip()
            try:
                data = pd.read_csv(csv_file)
            except Exception:
                continue
            if "Name" not in data.columns:
                continue
            for _, row in data.iterrows():
                name = str(row.get("Name", "")).strip()
                if not name:
                    continue
                time_str = str(row.get("Time", "")).strip()
                if time_str and time_str.lower() != "nan":
                    mark = f"P ({time_str})"
                else:
                    mark = "P"
                marks[(name, date_part, period)] = mark
        return marks

    def _write_class_month_workbook(self, output_path, rows, date_periods):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"

        fixed_headers = ["Batch ( year )", "Class name", "Semester", "Roll no.", "Name", "contact no."]
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="F3F6F4")
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_index, header in enumerate(fixed_headers, start=1):
            sheet.merge_cells(start_row=1, start_column=col_index, end_row=2, end_column=col_index)
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        start_col = len(fixed_headers) + 1
        for offset, item in enumerate(date_periods):
            col_index = start_col + offset
            top = sheet.cell(row=1, column=col_index, value=f"date ( dd-mm-yyyy ) Example - {item['display_date']}")
            bottom = sheet.cell(row=2, column=col_index, value=item["subject"] or f"P{item['period']}")
            for cell in (top, bottom):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border

        for row_index, row in enumerate(rows, start=3):
            values = [row[h] for h in fixed_headers] + row["marks"]
            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.alignment = center
                cell.border = border

        max_col = len(fixed_headers) + len(date_periods)
        for col_index in range(1, max_col + 1):
            letter = get_column_letter(col_index)
            if col_index == 5:
                sheet.column_dimensions[letter].width = 26
            elif col_index <= len(fixed_headers):
                sheet.column_dimensions[letter].width = 16
            else:
                sheet.column_dimensions[letter].width = 18

        sheet.freeze_panes = "G3"
        workbook.save(output_path)

    def _load_timetable(self, class_name):
        timetable_file = Path(TIMETABLE_DIR) / f"{class_name}.json"
        if not timetable_file.exists():
            return {"days": {}}
        try:
            with open(timetable_file, "r", encoding="utf-8") as handle:
                return json.load(handle)
        except Exception:
            return {"days": {}}

    def _get_period_details_for_date(self, cls, date_str, timetable):
        periods = {}
        try:
            day_name = datetime.strptime(date_str, "%Y-%m-%d").strftime("%A")
        except Exception:
            day_name = ""

        for item in timetable.get("days", {}).get(day_name, []):
            period = str(item.get("period", "")).strip()
            if period:
                periods[period] = {
                    "period": period,
                    "subject": item.get("subject", ""),
                }

        class_path = Path(self.base_dir) / cls
        if class_path.exists():
            for csv_file in sorted(class_path.glob(f"{date_str}_P*.csv")):
                period = csv_file.stem.split("_P")[-1].strip()
                if period and period not in periods:
                    periods[period] = {"period": period, "subject": f"P{period}"}

        if not periods:
            periods["1"] = {"period": "1", "subject": "P1"}

        return [periods[key] for key in sorted(periods, key=lambda value: int(value) if value.isdigit() else 9999)]

    def _split_class_and_semester(self, class_name):
        text = str(class_name or "").strip()
        match = re.search(r"\b(\d+)\s*(?:sem|semester|st|nd|rd|th)\b", text, flags=re.IGNORECASE)
        if not match:
            return text, "Semester"

        number = int(match.group(1))
        suffix = "th"
        if number % 100 not in (11, 12, 13):
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(number % 10, "th")
        semester = f"{number}{suffix}"
        clean_class = (text[: match.start()] + text[match.end() :]).strip(" -_")
        return clean_class or text, semester

    def _safe_filename(self, value):
        safe = re.sub(r'[<>:"/\\|?*\s]+', "_", str(value or "").strip())
        return safe.strip("_") or "Attendance"

    def _load_monthly_data(self, class_path, cls):
        groups = {}

        for file_name in sorted(os.listdir(class_path)):
            if not file_name.endswith(".csv") or "_P" not in file_name:
                continue

            full = os.path.join(class_path, file_name)
            try:
                day_df = pd.read_csv(full)
            except Exception:
                continue

            if day_df.empty:
                continue

            date_part = file_name.split("_P")[0]
            try:
                date_obj = datetime.strptime(date_part, "%Y-%m-%d")
            except ValueError:
                continue

            period_str = file_name.split("_P")[-1].replace(".csv", "").strip()
            month_key = date_obj.strftime("%Y-%m")

            if "Name" not in day_df.columns:
                continue

            if "Class" not in day_df.columns:
                day_df["Class"] = cls

            if "Time" not in day_df.columns:
                day_df["Time"] = ""

            day_df["Date"] = date_obj.strftime("%Y-%m-%d")
            day_df["Period"] = str(period_str)
            day_df["Time"] = day_df["Time"].astype(str)

            groups.setdefault(month_key, []).append(day_df[["Name", "Class", "Time", "Date", "Period"]])

        for key in list(groups.keys()):
            groups[key] = pd.concat(groups[key], ignore_index=True)

        return groups

    def _get_all_students_for_class(self, class_name, month_df):
        names = set()
        class_face_dir = os.path.join(self.face_dir, class_name)
        if os.path.exists(class_face_dir):
            for file_name in os.listdir(class_face_dir):
                if file_name.endswith(".npy"):
                    names.add(os.path.splitext(file_name)[0])

        if not month_df.empty and "Name" in month_df.columns:
            for n in month_df["Name"].dropna().astype(str):
                if n.strip():
                    names.add(n.strip())

        if not names:
            return pd.DataFrame(columns=["Name", "Class"])

        return pd.DataFrame(
            [{"Name": n, "Class": class_name} for n in sorted(names)],
            columns=["Name", "Class"],
        )

    def _get_timetable_periods(self, class_name):
        timetable_file = os.path.join(str(TIMETABLE_DIR), f"{class_name}.json")
        periods = set()

        if os.path.exists(timetable_file):
            try:
                with open(timetable_file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                for _, day_periods in data.get("days", {}).items():
                    for item in day_periods:
                        p = item.get("period")
                        if p is not None:
                            periods.add(str(p))
            except Exception:
                pass

        if not periods:
            periods = {"1"}

        return sorted(periods, key=lambda x: int(x) if str(x).isdigit() else 9999)

    def _build_month_matrix(self, month_df, month_key, periods):
        year, month = [int(x) for x in month_key.split("-")]
        days_in_month = calendar.monthrange(year, month)[1]
        all_dates = [f"{year:04d}-{month:02d}-{d:02d}" for d in range(1, days_in_month + 1)]

        students = month_df[["Name", "Class"]].dropna().drop_duplicates().sort_values(["Name", "Class"])
        if students.empty:
            return pd.DataFrame()

        idx = pd.MultiIndex.from_frame(students, names=["Name", "Class"])

        col_tuples = []
        for date_str in all_dates:
            for p in periods:
                col_tuples.append((date_str, f"P{p}"))

        cols = pd.MultiIndex.from_tuples(col_tuples, names=["Date", "Period"])
        report = pd.DataFrame("", index=idx, columns=cols)
        today = datetime.now().date()

        for date_str in all_dates:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
            for p in periods:
                col_key = (date_str, f"P{p}")
                if d > today:
                    base_val = ""
                elif d.weekday() == 6:
                    base_val = "H"
                else:
                    base_val = "A"
                report[col_key] = base_val

        for _, row in month_df.iterrows():
            name = str(row.get("Name", "")).strip()
            cls = str(row.get("Class", "")).strip()
            date_str = str(row.get("Date", "")).strip()
            period = str(row.get("Period", "")).strip()
            time_str = str(row.get("Time", "")).strip()

            if not name or not cls or not date_str or not period:
                continue
            if (name, cls) not in report.index:
                continue

            col_key = (date_str, f"P{period}")
            if col_key not in report.columns:
                continue

            mark = f"P ({time_str})" if time_str and time_str.lower() != "nan" else "P"
            report.at[(name, cls), col_key] = mark

        return report

    def _ensure_all_students(self, report_df, students_df):
        if students_df.empty:
            return report_df

        if report_df.empty:
            idx = pd.MultiIndex.from_frame(students_df, names=["Name", "Class"])
            return pd.DataFrame(index=idx)

        existing = set(report_df.index.tolist())
        to_add = []
        for _, row in students_df.iterrows():
            key = (str(row["Name"]), str(row["Class"]))
            if key not in existing:
                to_add.append(key)

        if not to_add:
            return report_df

        add_df = pd.DataFrame(
            [["" for _ in range(len(report_df.columns))] for _ in range(len(to_add))],
            index=pd.MultiIndex.from_tuples(to_add, names=report_df.index.names),
            columns=report_df.columns,
        )

        today = datetime.now().date()
        for date_str, period in report_df.columns:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
            if d > today:
                base_val = ""
            elif d.weekday() == 6:
                base_val = "H"
            else:
                base_val = "A"
            add_df[(date_str, period)] = base_val

        out = pd.concat([report_df, add_df], axis=0)
        out = out.sort_index()
        return out

    def showEvent(self, event):
        self.load_classes()
        super().showEvent(event)
